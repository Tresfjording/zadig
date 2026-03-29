from openpyxl import load_workbook
from openpyxl.utils import get_column_letter
from pathlib import Path
import locale

TXT_FIL = "ordliste_fra_sqlite.txt"
EXCEL_FIL = "Ordliste_Norsk_ny.xlsx"
ARK = "Ordliste"
KOLONNER = 29  # A:AC

# Norsk sortering
try:
    locale.setlocale(locale.LC_COLLATE, "nb_NO.UTF-8")
except locale.Error:
    locale.setlocale(locale.LC_COLLATE, "")

def norsk_sortering(ordliste):
    return sorted(ordliste, key=locale.strxfrm)

# Les alle ord
with open(TXT_FIL, encoding="utf-8") as f:
    ordliste = [linje.strip() for linje in f if linje.strip()]

ordliste = norsk_sortering(ordliste)

# Åpne Excel
wb = load_workbook(EXCEL_FIL, keep_vba=True)
ws = wb[ARK]

# Skriv ut i rader, 29 ord per rad (A:AC)
rad = 1
for i in range(0, len(ordliste), KOLONNER):
    chunk = ordliste[i:i+KOLONNER]
    for j, ord in enumerate(chunk):
        ws.cell(row=rad, column=j+1, value=ord)
    rad += 1

wb.save(EXCEL_FIL)
print(f"Ferdig! {len(ordliste)} ord importert til {EXCEL_FIL} [{ARK}!A:AC]")
