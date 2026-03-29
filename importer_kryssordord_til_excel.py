
import argparse
from openpyxl import load_workbook
from openpyxl.utils import column_index_from_string
from pathlib import Path

# Legg til kommandolinje-argument for kryssordfil
parser = argparse.ArgumentParser(description="Importer kryssordord til Excel")
parser.add_argument("--kryssordfil", type=str, default="kryssord_losningsord.txt", help="Sti til kryssordfil")
args = parser.parse_args()

TXT_FIL = args.kryssordfil
EXCEL_FIL = r"C:\Users\ØyvindGranberg\Documents\Ordliste\ordliste_norsk_ny.xlsx"
ARK = "Ordliste"
KOLONNE = "AG"  # Kolonne AG

def rens_ord(ordlinje):
    return ordlinje.replace(" ", "").replace("-", "").lower()

# Les og rens løsningsord
nye_ord = set()
with open(TXT_FIL, encoding="utf-8") as f:
    for linje in f:
        linje = linje.strip()
        if linje:
            nye_ord.add(rens_ord(linje))

# Åpne Excel og les eksisterende ord i valgt kolonne
try:
    wb = load_workbook(EXCEL_FIL)
except Exception as e:
    print(f"FEIL: Klarte ikke å åpne {EXCEL_FIL}: {e}")
    exit(1)

if ARK not in wb.sheetnames:
    print(f"FEIL: Fant ikke ark '{ARK}' i filen. Tilgjengelige ark: {wb.sheetnames}")
    exit(1)
ws = wb[ARK]

try:
    kol_idx = column_index_from_string(KOLONNE)
except Exception as e:
    print(f"FEIL: Ugyldig kolonnenavn '{KOLONNE}': {e}")
    exit(1)

eksisterende_ord = set()
for row in ws.iter_rows(min_col=kol_idx, max_col=kol_idx, min_row=1, values_only=True):
    verdi = row[0]
    if verdi:
        eksisterende_ord.add(rens_ord(str(verdi)))

# Finn bare nye ord som ikke allerede finnes
nye_unike = sorted(nye_ord - eksisterende_ord)
print(f"Antall nye ord som legges til: {len(nye_unike)}")

# Finn første ledige rad i valgt kolonne
rad = 1
while ws.cell(row=rad, column=kol_idx).value:
    rad += 1
print(f"Første ledige rad i kolonne {KOLONNE}: {rad}")

# Skriv inn nye ord og logg hvor det skrives
for ord in nye_unike:
    print(f"Skriver '{ord}' til rad {rad}, kolonne {KOLONNE} (indeks {kol_idx})")
    ws.cell(row=rad, column=kol_idx, value=ord)
    rad += 1

wb.save(EXCEL_FIL)
print(f"Ferdig! {len(nye_unike)} ord lagt til i {EXCEL_FIL} [{ARK}!{KOLONNE}]")
