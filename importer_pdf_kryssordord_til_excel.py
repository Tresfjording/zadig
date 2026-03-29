import os
import pdfplumber
from openpyxl import load_workbook
from openpyxl.utils import column_index_from_string

def rens_ord(ordlinje):
    return ordlinje.replace(" ", "").replace("-", "").lower()


# Sett PDF-mappe til F:\skannes
PDF_MAPPE = r"F:\skannes"
EXCEL_FIL = r"C:\Users\ØyvindGranberg\Documents\Ordliste\ordliste_norsk_ny.xlsx"
ARK = "Ordliste"
KOLONNE = "AG"  # Kolonne AG



# Les og hent løsningsord fra ALLE PDF-filer i mappen
resultat_ord = []  # Lagrer alle løsningsord (også tomme)
for filnavn in os.listdir(PDF_MAPPE):
    if filnavn.lower().endswith(".pdf"):
        pdf_path = os.path.join(PDF_MAPPE, filnavn)
        print(f"Leser PDF: {pdf_path}")
        try:
            with pdfplumber.open(pdf_path) as pdf:
                for side in pdf.pages:
                    tekst = side.extract_text()
                    if not tekst:
                        continue
                    linjer = [l.strip() for l in tekst.splitlines()]
                    i = 0
                    while i < len(linjer):
                        linje = linjer[i]
                        renset = rens_ord(linje)
                        # Oppgaveordrute: flere enn én bokstav
                        if len(renset) > 1 and renset.isalpha():
                            # Start å samle løsningsordruter
                            loesningsord = ""
                            j = i + 1
                            while j < len(linjer):
                                neste = rens_ord(linjer[j])
                                if not neste:
                                    break  # tom linje
                                if len(neste) > 1 and neste.isalpha():
                                    break  # ny oppgaveordrute
                                if len(neste) == 1 and neste.isalpha():
                                    loesningsord += neste
                                j += 1
                            resultat_ord.append(loesningsord)
                            i = j
                        else:
                            i += 1
        except Exception as e:
            print(f"FEIL ved lesing av {pdf_path}: {e}")


# Åpne Excel og les eksisterende ord i valgt kolonne
try:
    wb = load_workbook(EXCEL_FIL)
except Exception as e:
    print(f"FEIL: Klarte ikke å åpne {EXCEL_FIL}: {e}")
    exit(1)

if ARK not in wb.sheetnames:
    print(f"FEIL: Fant ikke ark '{ARK}' i filen. Tilgjengelige ark: {wb.sheetnames}")
    exit(1)
ws = wb[ARK]

try:
    kol_idx = column_index_from_string(KOLONNE)
except Exception as e:
    print(f"FEIL: Ugyldig kolonnenavn '{KOLONNE}': {e}")
    exit(1)

eksisterende_ord = set()
for row in ws.iter_rows(min_col=kol_idx, max_col=kol_idx, min_row=1, values_only=True):
    verdi = row[0]
    if verdi:
        eksisterende_ord.add(rens_ord(str(verdi)))

# Fjern duplikater og ord som allerede finnes, men behold tomme løsningsord
unik_ord = []
sett_ord = set(eksisterende_ord)
for ord in resultat_ord:
    if ord == "":
        unik_ord.append("")
    elif ord not in sett_ord:
        unik_ord.append(ord)
        sett_ord.add(ord)
print(f"Antall nye ord som legges til: {len([o for o in unik_ord if o])}")

# Finn første ledige rad i valgt kolonne
rad = 1
while ws.cell(row=rad, column=kol_idx).value:
    rad += 1
print(f"Første ledige rad i kolonne {KOLONNE}: {rad}")

# Skriv inn nye ord og logg hvor det skrives
for ord in unik_ord:
    if ord:
        print(f"Skriver '{ord}' til rad {rad}, kolonne {KOLONNE} (indeks {kol_idx})")
    else:
        print(f"Setter linjeskift på rad {rad}")
    ws.cell(row=rad, column=kol_idx, value=ord if ord else None)
    rad += 1

wb.save(EXCEL_FIL)
print(f"Ferdig! {len([o for o in unik_ord if o])} ord lagt til i {EXCEL_FIL} [{ARK}!{KOLONNE}]")
