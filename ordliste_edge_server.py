from __future__ import annotations

import html
import os
import sqlite3
import threading
import time
import webbrowser
from dataclasses import dataclass
from http import HTTPStatus
from http.server import BaseHTTPRequestHandler, HTTPServer
from pathlib import Path
from typing import Iterable
from urllib.parse import parse_qs, urlparse

from openpyxl import load_workbook
from openpyxl.utils import column_index_from_string


ORDLISTE_SHEET_CANDIDATES = ["Ordliste", "AlleOrd", "Kolonner"]
SEARCH_SHEET_CANDIDATES = ["SearchWords", "test"]
SEARCH_CELL = "D3"
COLUMN_START = "A"
COLUMN_END = "AC"
DB_PATH = Path(__file__).with_name("ordliste_cache.sqlite3")
HOST = "127.0.0.1"
PORT = 8765
UI_VERSION = "v2026-03-26-wildcards-limitfix"
DEFAULT_SAMPLE_LIMIT = 0

def do_GET(self) -> None:
    self.send_response(200)
    self.send_header("Content-Type", "text/html; charset=utf-8")
    self.end_headers()
    self.wfile.write(b"<h1>Serveren svarer!</h1>")

def resolve_excel_path() -> Path:
    env_path = os.environ.get("ORDLISTE_XLSM")
    candidates = [
        Path(env_path) if env_path else None,
        # Prioriter ny hovedfil først:
        Path.cwd() / "Ordliste_Norsk_ny.xlsx",
        # Deretter gamle filer om ny ikke finnes:
        Path.cwd() / "G-Ordliste.xlsm",
        # Path.cwd() / "Ordliste Norsk v. 30.6.xlsm",
        # Path.cwd() / "Ordlista HovedFil.xlsm",
        # Path.cwd() / "Ordliste Norsk.xlsm",
        # Originale OneDrive-steder
        # Path(r"C:\Users\Øyvind – ZaGal ote\OneDrive\Dokumenter\Annet\Ordlista HovedFil.xlsm"),
        # Path(r"C:\Users\Øyvind – ZaGal ote\OneDrive\Dokumenter\Annet\Ordliste Norsk.xlsm"),
    ]

    valid_candidates = [path for path in candidates if path is not None]
    for candidate in valid_candidates:
        if candidate.is_file():
            return candidate

    # Fallback to the first candidate so the error message points til en konkret path.
    return valid_candidates[0]


EXCEL_PATH = resolve_excel_path()


def normalize_word(value: object) -> str:
    if value is None:
        return ""
    text = str(value).strip().lower()
    return " ".join(text.split())


def wildcard_to_sql_like(pattern: str) -> str:
    translated: list[str] = []
    for char in pattern:
        if char == "*":
            translated.append("%")
        elif char == "?":
            translated.append("_")
        elif char in {"%", "_", "\\"}:
            translated.append("\\" + char)
        else:
            translated.append(char)
    return "".join(translated)


def find_edge_executable() -> Path | None:
    candidates = [
        Path(os.environ.get("ProgramFiles(x86)", "")) / "Microsoft" / "Edge" / "Application" / "msedge.exe",
        Path(os.environ.get("ProgramFiles", "")) / "Microsoft" / "Edge" / "Application" / "msedge.exe",
        Path(os.environ.get("LocalAppData", "")) / "Microsoft" / "Edge" / "Application" / "msedge.exe",
    ]
    for candidate in candidates:
        if candidate.is_file():
            return candidate
    return None


@dataclass
class SearchResult:
    query: str
    normalized_query: str
    found: bool
    wildcard_used: bool
    match_count: int
    sample_matches: list[str]
    sample_limit: int
    total_words: int
    source: str
    indexed_at: str | None
    message: str


class WordIndex:
    def __init__(self, excel_path: Path, db_path: Path) -> None:
        self.excel_path = excel_path
        self.db_path = db_path
        self._lock = threading.Lock()
        self.active_word_sheet: str | None = None
        self.active_search_sheet: str | None = None
        self._ensure_schema()

    def _connect(self) -> sqlite3.Connection:
        connection = sqlite3.connect(self.db_path)
        connection.execute("PRAGMA journal_mode=WAL")
        connection.execute("PRAGMA synchronous=NORMAL")
        return connection

    def _ensure_schema(self) -> None:
        with self._connect() as connection:
            connection.execute(
                """
                CREATE TABLE IF NOT EXISTS metadata (
                    key TEXT PRIMARY KEY,
                    value TEXT NOT NULL
                )
                """
            )
            connection.execute(
                """
                CREATE TABLE IF NOT EXISTS words (
                    word TEXT PRIMARY KEY
                )
                """
            )
            connection.execute("CREATE INDEX IF NOT EXISTS idx_words_word ON words(word)")

    def _get_metadata(self, key: str) -> str | None:
        with self._connect() as connection:
            row = connection.execute("SELECT value FROM metadata WHERE key = ?", (key,)).fetchone()
        return row[0] if row else None

    def _set_metadata(self, key: str, value: str) -> None:
        with self._connect() as connection:
            connection.execute(
                "INSERT INTO metadata(key, value) VALUES (?, ?) "
                "ON CONFLICT(key) DO UPDATE SET value = excluded.value",
                (key, value),
            )

    @staticmethod
    def _get_sheet_from_candidates_or_raise(workbook, candidates: list[str], label: str):
        normalized_map = {name.strip().casefold(): name for name in workbook.sheetnames}
        for candidate in candidates:
            key = candidate.strip().casefold()
            if key in normalized_map:
                return workbook[normalized_map[key]], normalized_map[key]

        available = ", ".join(workbook.sheetnames)
        wanted = ", ".join(candidates)
        raise KeyError(
            f"Fant ikke {label}. Prøvde: {wanted}. Tilgjengelige ark: {available}"
        )

    def _iter_words_from_workbook(self) -> Iterable[str]:
        workbook = load_workbook(
            self.excel_path,
            read_only=True,
            data_only=True,
            keep_vba=True,
            keep_links=False,
        )
        try:
            sheet, actual_name = self._get_sheet_from_candidates_or_raise(
                workbook,
                ORDLISTE_SHEET_CANDIDATES,
                "ordliste-ark",
            )
            self.active_word_sheet = actual_name
            min_column = column_index_from_string(COLUMN_START)
            max_column = column_index_from_string(COLUMN_END)
            for row_values in sheet.iter_rows(min_col=min_column, max_col=max_column, values_only=True):
                for cell_value in row_values:
                    word = normalize_word(cell_value)
                    if word:
                        yield word
        finally:
            workbook.close()

    def rebuild_if_needed(self) -> bool:
        if not self.excel_path.is_file():
            raise FileNotFoundError(f"Fant ikke Excel-filen: {self.excel_path}")

        stat = self.excel_path.stat()
        signature = f"{stat.st_mtime_ns}:{stat.st_size}"
        cached_signature = self._get_metadata("excel_signature")
        if cached_signature == signature:
            return False

        with self._lock:
            cached_signature = self._get_metadata("excel_signature")
            if cached_signature == signature:
                return False

            print("Bygger indeks fra Excel ...")
            start_time = time.perf_counter()
            indexed_at = time.strftime("%Y-%m-%d %H:%M:%S")
            scanned = 0
            batch: list[tuple[str]] = []
            batch_size = 10000

            with self._connect() as connection:
                connection.execute("DELETE FROM words")

                for word in self._iter_words_from_workbook():
                    scanned += 1
                    batch.append((word,))
                    if len(batch) >= batch_size:
                        connection.executemany("INSERT OR IGNORE INTO words(word) VALUES (?)", batch)
                        batch.clear()
                    if scanned % 250000 == 0:
                        print(f"  Leste {scanned:,} celler ...".replace(",", " "))

                if batch:
                    connection.executemany("INSERT OR IGNORE INTO words(word) VALUES (?)", batch)

                unique_count = connection.execute("SELECT COUNT(*) FROM words").fetchone()[0]
                connection.execute(
                    "INSERT INTO metadata(key, value) VALUES (?, ?) "
                    "ON CONFLICT(key) DO UPDATE SET value = excluded.value",
                    ("excel_signature", signature),
                )
                connection.execute(
                    "INSERT INTO metadata(key, value) VALUES (?, ?) "
                    "ON CONFLICT(key) DO UPDATE SET value = excluded.value",
                    ("indexed_at", indexed_at),
                )
                connection.execute(
                    "INSERT INTO metadata(key, value) VALUES (?, ?) "
                    "ON CONFLICT(key) DO UPDATE SET value = excluded.value",
                    ("word_count", str(unique_count)),
                )

            elapsed = time.perf_counter() - start_time
            print(
                (
                    f"Indeks ferdig. Leste {scanned:,} celler, "
                    f"{unique_count:,} unike ord, tid {elapsed:.1f}s"
                ).replace(",", " ")
            )

        return True

    def count_words(self) -> int:
        cached = self._get_metadata("word_count")
        return int(cached) if cached else 0

    def indexed_at(self) -> str | None:
        return self._get_metadata("indexed_at")

    def contains(self, query: str) -> bool:
        normalized = normalize_word(query)
        if not normalized:
            return False
        with self._connect() as connection:
            row = connection.execute("SELECT 1 FROM words WHERE word = ? LIMIT 1", (normalized,)).fetchone()
        return row is not None

    def find_wildcard_matches(self, query: str, limit: int = 50) -> tuple[int, list[str]]:
        normalized = normalize_word(query)
        if not normalized:
            return 0, []

        like_pattern = wildcard_to_sql_like(normalized)
        with self._connect() as connection:
            count_row = connection.execute(
                "SELECT COUNT(*) FROM words WHERE word LIKE ? ESCAPE '\\'",
                (like_pattern,),
            ).fetchone()
            if limit <= 0:
                sample_rows = connection.execute(
                    "SELECT word FROM words WHERE word LIKE ? ESCAPE '\\' ORDER BY word",
                    (like_pattern,),
                ).fetchall()
            else:
                sample_rows = connection.execute(
                    "SELECT word FROM words WHERE word LIKE ? ESCAPE '\\' ORDER BY word LIMIT ?",
                    (like_pattern, limit),
                ).fetchall()

        total = int(count_row[0]) if count_row else 0
        samples = [row[0] for row in sample_rows]
        return total, samples

    def read_search_cell(self) -> str:
        workbook = load_workbook(
            self.excel_path,
            read_only=True,
            data_only=True,
            keep_vba=True,
            keep_links=False,
        )
        try:
            sheet, actual_name = self._get_sheet_from_candidates_or_raise(
                workbook,
                SEARCH_SHEET_CANDIDATES,
                "søkeark",
            )
            self.active_search_sheet = actual_name
            value = sheet[SEARCH_CELL].value
            return "" if value is None else str(value)
        finally:
            workbook.close()


class SearchApp:
    def make_result(self, query: str, source: str, sample_limit: int = DEFAULT_SAMPLE_LIMIT) -> SearchResult:
        normalized_query = normalize_word(query)
        wildcard_used = '*' in normalized_query or '?' in normalized_query
        found = False
        match_count = 0
        sample_matches = []
        message = ""
        if normalized_query:
            match_count, sample_matches = self.index.find_wildcard_matches(normalized_query, limit=sample_limit)
            found = match_count > 0
            if not found:
                message = f"Fant ingen treff for '{query}'."
        else:
            message = "Skriv inn et søkeord. Bruk * og ? for jokertegn om ønskelig."
        return SearchResult(
            query=query,
            normalized_query=normalized_query,
            found=found,
            wildcard_used=wildcard_used,
            match_count=match_count,
            sample_matches=sample_matches,
            sample_limit=sample_limit,
            total_words=self.index.count_words(),
            source=source,
            indexed_at=self.index.indexed_at(),
            message=message,
        )
    def __init__(self, index: WordIndex):
        self.index = index

    def result_from_sheet_or_error(self, sample_limit=DEFAULT_SAMPLE_LIMIT):
        try:
            query = self.index.read_search_cell()
            return self.make_result(query, "Excel SearchWords!D3", sample_limit=sample_limit)
        except Exception as exc:
            return SearchResult(
    query="",
    normalized_query="",
    found=False,
    wildcard_used=False,
    match_count=0,
    sample_matches=[],
    sample_limit=sample_limit,
    total_words=self.index.count_words(),
    source="sheet",
    indexed_at=self.index.indexed_at(),
    message="",  # <-- behold denne, evt. tom streng
)
    def render_page(self, result: SearchResult) -> str:
        safe_query = html.escape(result.query or "")
        safe_samples = "".join(f"<li>{html.escape(word)}</li>" for word in (result.sample_matches or []))
        samples_style = "" if result.found else "color: #7c4a03;"
        sample_limit = "alle" if result.sample_limit == 0 else f"{result.sample_limit:,}".replace(",", " ")
        # Vis kun feilmelding hvis source er 'sheet' og det faktisk er en feilmelding
        show_message = bool(result.message and result.source == "sheet")
        message = html.escape(result.message) if show_message else ""
        return f"""<!doctype html>
<html lang=\"no\">
<head>
    <meta charset=\"utf-8\">
    <title>Ordliste-søk</title>
    <style>
        body {{ font-family: Arial, sans-serif; background: #fff8e1; margin: 0; padding: 0; }}
        .card {{ background: #fff3e0; max-width: 600px; margin: 40px auto; padding: 24px 32px; border-radius: 12px; box-shadow: 0 2px 8px rgba(140, 70, 0, 0.10); }}
        .hero {{ margin-bottom: 18px; }}
        .actions {{ display: flex; gap: 10px; margin-top: 12px; margin-bottom: 12px; }}
        .button-link, button {{ border: none; border-radius: 999px; padding: 14px 18px; font-size: 0.95rem; cursor: pointer; text-decoration: none; color: #fffbe6; background: #e65100; transition: background 0.2s; }}
        .button-link:hover, button:hover {{ background: #ff9800; color: #4e2600; }}
        .button-secondary {{ background: #a1887f; color: #fffbe6; }}
        .button-secondary:hover {{ background: #bcaaa4; color: #4e2600; }}
        .message {{ font-size: 1rem; line-height: 1.5; color: #b71c1c; margin-bottom: 10px; }}
        .samples-block {{ margin-top: 6px; border-top: 1px solid #bcaaa4; padding-top: 14px; }}
        .samples-title {{ margin: 0 0 8px; font-size: 0.92rem; color: #a1887f; letter-spacing: 0.02em; }}
        .samples {{ margin: 0; padding-left: 20px; {samples_style} }}
        .samples li {{ margin-bottom: 2px; }}
        .footer {{ font-size: 0.88rem; color: #a1887f; margin-top: 18px; }}
        .match-count {{ font-size: 1.05rem; color: #e65100; font-weight: bold; margin-bottom: 6px; }}
        input[type="text"], input[type="number"] {{ border: 1px solid #ff9800; border-radius: 6px; padding: 6px 10px; background: #fffbe6; color: #4e2600; margin-right: 8px; }}
        input[type="text"]:focus, input[type="number"]:focus {{ outline: 2px solid #e65100; }}
        h1 {{ color: #e65100; }}
        label {{ color: #7c4a03; }}
        @media (max-width: 640px) {{ body {{ padding: 16px; }} .hero, .card {{ padding-left: 10px; padding-right: 10px; }} .actions {{ flex-direction: column; }} button, .button-link {{ width: 100%; text-align: center; }} }}
    </style>
</head>
<body>
    <main class=\"card\">
        <section class=\"hero\">
            <h1>Ordliste-søk</h1>
            <form method=\"get\" action=\"/\">
                <label for=\"term\">Søk etter ord:</label>
                <input type=\"text\" id=\"term\" name=\"term\" value=\"{safe_query}\" autocomplete=\"off\" autofocus>
                <label for=\"limit\">Antall eksempler:</label>
                <input type=\"number\" id=\"limit\" name=\"limit\" value=\"{result.sample_limit}\" min=\"0\" max=\"1000\">
                <div class=\"actions\">
                    <button type=\"submit\">Søk i ordlisten</button>
                    <a class=\"button-link button-secondary\" href=\"/?source=sheet&limit={DEFAULT_SAMPLE_LIMIT}\">Les SearchWords!D3</a>
                    <a class=\"button-link\" href=\"/rebuild?limit={DEFAULT_SAMPLE_LIMIT}\">Bygg indeks på nytt</a>
                </div>
                <div class=\"samples-block\">
                    <div class=\"match-count\">Antall treff: {result.match_count:,}</div>
                    <p class=\"samples-title\">Treffliste (viser inntil {result.total_words} ord)</p>
                    <ul class=\"samples\">{safe_samples or '<li>Ingen treff å vise</li>'}</ul>
                </div>
            </form>
            {f'<div class="message">{message}</div>' if message else ''}
            <div class=\"footer\">Excel-fil: {html.escape(str(EXCEL_PATH))} </div>
        </section>
    </main>
</body>
</html>"""
        #sample_limit = "alle" if result.sample_limit == 0 else f"{result.sample_limit:,}".replace(",", " ")
        return f"""<!doctype html>
<html lang=\"no\">
"""
def result_from_sheet_or_error(self, sample_limit=DEFAULT_SAMPLE_LIMIT):
        try:
            query = self.index.read_search_cell()
            return self.make_result(query, "Excel SearchWords!D3", sample_limit=sample_limit)
        except Exception as exc:
            return SearchResult(
                query="",
                normalized_query="",
                found=False,
                wildcard_used=False,
                match_count=0,
                sample_matches=[],
                sample_limit=sample_limit,
                total_words=self.index.count_words(),
                source="sheet",
                indexed_at=self.index.indexed_at(),
                message=f"Klarte ikke lese fra SearchWords!D3: {exc}",
            )
def make_handler(app: SearchApp) -> type[BaseHTTPRequestHandler]:


    class Handler(BaseHTTPRequestHandler):
        def do_GET(self) -> None:
            print(f"do_GET: Forespørsel mottatt: {self.path}")
            try:
                app = self.server.app  # type: ignore[attr-defined]
                print("do_GET: Parsed app OK")
                from urllib.parse import urlparse, parse_qs
                parsed = urlparse(self.path)
                print(f"do_GET: Parsed URL: {parsed}")
                params = parse_qs(parsed.query)
                print(f"do_GET: Params: {params}")
                sample_limit = 0
                try:
                    if "limit" in params:
                        sample_limit = int(params["limit"][0])
                except Exception as e:
                    print(f"do_GET: Feil ved parsing av limit: {e}")
                    sample_limit = 0

                if parsed.path == "/rebuild":
                    print("do_GET: Rebuild path")
                    try:
                        app.index.rebuild_from_excel()
                        message = "Indeksen ble bygget på nytt fra Excel."
                    except Exception as exc:
                        print(f"do_GET: Feil ved rebuild: {exc}")
                        result = SearchResult(
                            query="",
                            normalized_query="",
                            found=False,
                            wildcard_used=False,
                            match_count=0,
                            sample_matches=[],
                            sample_limit=sample_limit,
                            total_words=0,
                            source="rebuild",
                            indexed_at=None,
                            message=f"Klarte ikke bygge indeksen på nytt: {exc}",
                        )
                        self._send_html(app.render_page(result))
                        return
                    result = app.make_result("", "rebuild", sample_limit=sample_limit)
                    self._send_html(app.render_page(result))
                    return

                source = params.get("source", [""])[0]
                print(f"do_GET: source={source}")
                if source == "sheet":
                    print("do_GET: Henter fra sheet")
                    result = app.result_from_sheet_or_error(sample_limit=sample_limit)
                else:
                    query = params.get("term", [""])[0]
                    print(f"do_GET: Henter fra manuell input, query={query}")
                    result = app.make_result(query, "manuell input", sample_limit=sample_limit)
                print("do_GET: Sender HTML-svar")
                self._send_html(app.render_page(result))
            except Exception as e:
                print(f"do_GET: Uventet feil: {e}")
                self.send_response(500)
                self.send_header("Content-Type", "text/plain; charset=utf-8")
                self.end_headers()
                self.wfile.write(f"Serverfeil: {e}".encode("utf-8"))

        def log_message(self, format: str, *args: object) -> None:
            return

        def _send_html(self, body: str) -> None:
            encoded = body.encode("utf-8")
            self.send_response(200)
            self.send_header("Content-Type", "text/html; charset=utf-8")
            self.send_header("Content-Length", str(len(encoded)))
            self.end_headers()
            self.wfile.write(encoded)

    return Handler


def open_in_edge(url: str) -> None:
    edge_path = find_edge_executable()
    if edge_path is not None:
        webbrowser.register(
            "edge-local",
            None,
            webbrowser.BackgroundBrowser(str(edge_path)),
            preferred=True,
        )
        webbrowser.get("edge-local").open(url)
        return
    webbrowser.open(url)


def main() -> int:

    index = WordIndex(EXCEL_PATH, DB_PATH)
    app = SearchApp(index)
    handler = make_handler(app)
    server = HTTPServer((HOST, PORT), handler)
    server.app = app  # Attach the app to the server for handler access
    url = f"http://{HOST}:{PORT}/?source=sheet&limit={DEFAULT_SAMPLE_LIMIT}"

    print(f"Excel-fil i bruk: {EXCEL_PATH}")
    print("Sjekker om indeksen er oppdatert ...")
    try:
        if index.rebuild_if_needed():
            print("Indeksen ble oppdatert ved oppstart.")
        else:
            print("Indeksen er allerede oppdatert.")
    except Exception as exc:
        print(f"Kunne ikke forhåndsbygge indeks: {exc}")

    # Tell antall unike ord direkte fra Excel-arket (A:AC, Ordliste)
    try:
        from openpyxl import load_workbook
        from openpyxl.utils import column_index_from_string
        SHEET_CANDIDATES = ["Ordliste", "AlleOrd", "Kolonner"]
        COLUMN_START = "A"
        COLUMN_END = "AC"
        def normalize_word(value):
            if value is None:
                return ""
            text = str(value).strip().upper()
            return " ".join(text.split())
        wb = load_workbook(EXCEL_PATH, read_only=True, data_only=True, keep_vba=True, keep_links=False)
        try:
            normalized_map = {name.strip().casefold(): name for name in wb.sheetnames}
            for candidate in SHEET_CANDIDATES:
                key = candidate.strip().casefold()
                if key in normalized_map:
                    sheet = wb[normalized_map[key]]
                    break
            else:
                raise KeyError(f"Fant ikke ordliste-ark. Prøvde: {SHEET_CANDIDATES}. Tilgjengelige ark: {wb.sheetnames}")
            min_col = column_index_from_string(COLUMN_START)
            max_col = column_index_from_string(COLUMN_END)
            words = set()
            for row in sheet.iter_rows(min_col=min_col, max_col=max_col, values_only=True):
                for value in row:
                    word = normalize_word(value)
                    if word:
                        words.add(word)
            print(f"Antall unike ord i Excel-arket: {len(words):,}")
        finally:
            wb.close()
    except Exception as exc:
        print(f"Klarte ikke telle ord direkte fra Excel: {exc}")

    print(f"Antall ord i SQLite-indeksen: {index.count_words():,}")
    print(f"Starter lokal server på {url}")
    print("Trykk Ctrl+C for å stoppe serveren.")

    try:
        open_in_edge(url)
        server.serve_forever()
    except KeyboardInterrupt:
        print("Stopper serveren.")
    finally:
        server.server_close()

    # return 0

if __name__ == "__main__":
        raise SystemExit(main())