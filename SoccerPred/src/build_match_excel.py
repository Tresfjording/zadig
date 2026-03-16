import argparse
from pathlib import Path

import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Font
from openpyxl.worksheet.datavalidation import DataValidation


def parse_args() -> argparse.Namespace:
    parser = argparse.ArgumentParser(
        description="Lag Excel-ark for valg av hjemme- og bortelag med statistikk."
    )
    parser.add_argument(
        "--dataset",
        type=str,
        default="data/fifa_world_cup_top_scorers.csv",
        help="Sti til CSV-datasett.",
    )
    parser.add_argument(
        "--output",
        type=str,
        default="data/kampanalyse.xlsx",
        help="Output-fil for Excel-arket.",
    )
    parser.add_argument(
        "--matches-dataset",
        type=str,
        default="",
        help="Valgfri sti til kamp-CSV (for eksempel med HomeTeam/AwayTeam/HomeGoals/AwayGoals).",
    )
    return parser.parse_args()


def autosize_columns(ws, min_width: int = 10, max_width: int = 40) -> None:
    for col in ws.columns:
        max_len = 0
        for cell in col:
            if cell.value is None:
                continue
            max_len = max(max_len, len(str(cell.value)))
        ws.column_dimensions[col[0].column_letter].width = min(
            max(max_len + 2, min_width), max_width
        )


def build_team_stats(df: pd.DataFrame) -> pd.DataFrame:
    df = df.copy()
    df["Goals"] = pd.to_numeric(df["Goals"], errors="coerce").fillna(0)
    df["Year"] = pd.to_numeric(df["Year"], errors="coerce")

    stats = (
        df.groupby("Country", dropna=True)
        .agg(
            entries=("Player", "count"),
            unique_players=("Player", "nunique"),
            total_goals=("Goals", "sum"),
            avg_goals=("Goals", "mean"),
            max_goals=("Goals", "max"),
            first_year=("Year", "min"),
            last_year=("Year", "max"),
        )
        .reset_index()
        .sort_values("Country")
    )

    stats.rename(
        columns={
            "Country": "Team",
            "entries": "Entries",
            "unique_players": "UniquePlayers",
            "total_goals": "TotalGoals",
            "avg_goals": "AvgGoals",
            "max_goals": "MaxGoals",
            "first_year": "FirstYear",
            "last_year": "LastYear",
        },
        inplace=True,
    )

    return stats


def normalize_match_columns(df: pd.DataFrame) -> pd.DataFrame:
    rename_map = {}
    lower_to_real = {c.lower(): c for c in df.columns}

    candidates = {
        "Date": ["date", "match_date"],
        "HomeTeam": ["hometeam", "home_team", "home"],
        "AwayTeam": ["awayteam", "away_team", "away"],
        "HomeGoals": ["homegoals", "home_goals", "fthg", "home_score"],
        "AwayGoals": ["awaygoals", "away_goals", "ftag", "away_score"],
    }

    for target, keys in candidates.items():
        for key in keys:
            if key in lower_to_real:
                rename_map[lower_to_real[key]] = target
                break

    return df.rename(columns=rename_map)


def create_workbook(
    raw_df: pd.DataFrame,
    stats_df: pd.DataFrame,
    output_path: Path,
    matches_df: pd.DataFrame | None = None,
) -> None:
    wb = Workbook()

    ws_raw = wb.active
    ws_raw.title = "RawData"

    ws_stats = wb.create_sheet("TeamStats")
    ws_match = wb.create_sheet("Kampanalyse")
    ws_table = wb.create_sheet("Tabell")
    ws_live = wb.create_sheet("LiveTabell")
    ws_matches = wb.create_sheet("Kamper")

    # Raw data
    ws_raw.append(list(raw_df.columns))
    for row in raw_df.itertuples(index=False):
        ws_raw.append(list(row))

    # Team stats
    ws_stats.append(list(stats_df.columns))
    for row in stats_df.itertuples(index=False):
        ws_stats.append(list(row))

    # Match analysis UI
    ws_match["A1"] = "Kampanalyse"
    ws_match["A1"].font = Font(bold=True, size=14)

    ws_match["A2"] = "Hjemmelag"
    ws_match["A3"] = "Bortelag"

    ws_match["A5"] = "Statistikk"
    ws_match["B5"] = "Hjemme"
    ws_match["C5"] = "Borte"
    ws_match["D5"] = "Differanse"

    for cell in ("A5", "B5", "C5", "D5"):
        ws_match[cell].font = Font(bold=True)

    stat_labels = [
        "Entries",
        "UniquePlayers",
        "TotalGoals",
        "AvgGoals",
        "MaxGoals",
        "FirstYear",
        "LastYear",
    ]

    for i, label in enumerate(stat_labels, start=6):
        ws_match.cell(row=i, column=1, value=label)

    max_stats_row = ws_stats.max_row

    # Dropdown for team selection
    validation = DataValidation(
        type="list",
        formula1=f"=TeamStats!$A$2:$A${max_stats_row}",
        allow_blank=False,
    )
    ws_match.add_data_validation(validation)
    validation.add("B2")
    validation.add("B3")

    # Formulas for home/away stats and difference
    for i in range(6, 13):
        col_index = i - 4  # B=2 for row 6, C=3 ... in TeamStats
        ws_match.cell(
            row=i,
            column=2,
            value=(
                f"=IFERROR(INDEX(TeamStats!$B$2:$H${max_stats_row},"
                f"MATCH($B$2,TeamStats!$A$2:$A${max_stats_row},0),{col_index}),\"\")"
            ),
        )
        ws_match.cell(
            row=i,
            column=3,
            value=(
                f"=IFERROR(INDEX(TeamStats!$B$2:$H${max_stats_row},"
                f"MATCH($B$3,TeamStats!$A$2:$A${max_stats_row},0),{col_index}),\"\")"
            ),
        )
        ws_match.cell(row=i, column=4, value=f"=IF(OR(B{i}=\"\",C{i}=\"\"),\"\",B{i}-C{i})")

    ws_match["A4"] = "Tips"
    ws_match["B4"] = '=IF(B2=B3,"Velg to ulike lag","OK")'

    # Number formats
    ws_match["B9"].number_format = "0.00"
    ws_match["C9"].number_format = "0.00"
    ws_match["D9"].number_format = "0.00"

    normalized_matches = None
    teams = stats_df["Team"].dropna().tolist()
    if matches_df is not None and not matches_df.empty:
        candidate = normalize_match_columns(matches_df)
        required = {"HomeTeam", "AwayTeam", "HomeGoals", "AwayGoals"}
        if required.issubset(set(candidate.columns)):
            normalized_matches = candidate
            match_teams = pd.concat(
                [candidate["HomeTeam"], candidate["AwayTeam"]], ignore_index=True
            ).dropna()
            teams = sorted(match_teams.astype(str).unique().tolist())

    # Matches input/reference sheet
    match_headers = ["Date", "HomeTeam", "AwayTeam", "HomeGoals", "AwayGoals"]
    ws_matches.append(match_headers)
    for col in range(1, len(match_headers) + 1):
        ws_matches.cell(row=1, column=col).font = Font(bold=True)

    if normalized_matches is not None and not normalized_matches.empty:
        for _, row in normalized_matches.iterrows():
            ws_matches.append([
                row.get("Date", ""),
                row.get("HomeTeam", ""),
                row.get("AwayTeam", ""),
                row.get("HomeGoals", ""),
                row.get("AwayGoals", ""),
            ])
    elif matches_df is not None and not matches_df.empty:
        ws_matches["A2"] = "Fant kampfil, men mangler nødvendige kolonner."
        ws_matches["A3"] = "Trenger: HomeTeam, AwayTeam, HomeGoals, AwayGoals"
    else:
        ws_matches["A2"] = "Legg inn kamper her eller bruk --matches-dataset med CSV."

    # League table template (automatic from Kamper sheet)
    table_headers = ["Pos", "Team", "MP", "W", "D", "L", "GF", "GA", "Pts", "SortKey"]
    ws_table.append(table_headers)
    for col in range(1, len(table_headers) + 1):
        ws_table.cell(row=1, column=col).font = Font(bold=True)

    last_row = len(teams) + 1
    for idx, team in enumerate(teams, start=2):
        ws_table.cell(row=idx, column=2, value=team)
        ws_table.cell(
            row=idx,
            column=4,
            value=(
                f"=SUMPRODUCT((Kamper!$B$2:$B$5000=$B{idx})*(Kamper!$D$2:$D$5000>Kamper!$E$2:$E$5000))"
                f"+SUMPRODUCT((Kamper!$C$2:$C$5000=$B{idx})*(Kamper!$E$2:$E$5000>Kamper!$D$2:$D$5000))"
            ),
        )
        ws_table.cell(
            row=idx,
            column=5,
            value=(
                f"=SUMPRODUCT((Kamper!$B$2:$B$5000=$B{idx})*(Kamper!$D$2:$D$5000=Kamper!$E$2:$E$5000))"
                f"+SUMPRODUCT((Kamper!$C$2:$C$5000=$B{idx})*(Kamper!$E$2:$E$5000=Kamper!$D$2:$D$5000))"
            ),
        )
        ws_table.cell(
            row=idx,
            column=6,
            value=(
                f"=SUMPRODUCT((Kamper!$B$2:$B$5000=$B{idx})*(Kamper!$D$2:$D$5000<Kamper!$E$2:$E$5000))"
                f"+SUMPRODUCT((Kamper!$C$2:$C$5000=$B{idx})*(Kamper!$E$2:$E$5000<Kamper!$D$2:$D$5000))"
            ),
        )
        ws_table.cell(row=idx, column=3, value=f"=D{idx}+E{idx}+F{idx}")
        ws_table.cell(
            row=idx,
            column=7,
            value=f"=SUMIFS(Kamper!$D:$D,Kamper!$B:$B,$B{idx})+SUMIFS(Kamper!$E:$E,Kamper!$C:$C,$B{idx})",
        )
        ws_table.cell(
            row=idx,
            column=8,
            value=f"=SUMIFS(Kamper!$E:$E,Kamper!$B:$B,$B{idx})+SUMIFS(Kamper!$D:$D,Kamper!$C:$C,$B{idx})",
        )
        ws_table.cell(row=idx, column=9, value=f"=D{idx}*3+E{idx}")
        ws_table.cell(row=idx, column=10, value=f"=I{idx}*1000+(G{idx}-H{idx})")
        ws_table.cell(
            row=idx,
            column=1,
            value=f"=RANK(J{idx},$J$2:$J${last_row},0)+COUNTIF($J$2:J{idx},J{idx})-1",
        )

    ws_table.column_dimensions["J"].hidden = True

    # Live standings view (physically sorted by Pos)
    live_headers = ["Pos", "Team", "MP", "W", "D", "L", "GF", "GA", "Pts"]
    ws_live.append(live_headers)
    for col in range(1, len(live_headers) + 1):
        ws_live.cell(row=1, column=col).font = Font(bold=True)

    for row_idx in range(2, last_row + 1):
        ws_live.cell(row=row_idx, column=1, value=row_idx - 1)
        ws_live.cell(
            row=row_idx,
            column=2,
            value=f"=IFERROR(INDEX(Tabell!$B$2:$B${last_row},MATCH($A{row_idx},Tabell!$A$2:$A${last_row},0)),\"\")",
        )
        ws_live.cell(
            row=row_idx,
            column=3,
            value=f"=IFERROR(INDEX(Tabell!$C$2:$C${last_row},MATCH($A{row_idx},Tabell!$A$2:$A${last_row},0)),\"\")",
        )
        ws_live.cell(
            row=row_idx,
            column=4,
            value=f"=IFERROR(INDEX(Tabell!$D$2:$D${last_row},MATCH($A{row_idx},Tabell!$A$2:$A${last_row},0)),\"\")",
        )
        ws_live.cell(
            row=row_idx,
            column=5,
            value=f"=IFERROR(INDEX(Tabell!$E$2:$E${last_row},MATCH($A{row_idx},Tabell!$A$2:$A${last_row},0)),\"\")",
        )
        ws_live.cell(
            row=row_idx,
            column=6,
            value=f"=IFERROR(INDEX(Tabell!$F$2:$F${last_row},MATCH($A{row_idx},Tabell!$A$2:$A${last_row},0)),\"\")",
        )
        ws_live.cell(
            row=row_idx,
            column=7,
            value=f"=IFERROR(INDEX(Tabell!$G$2:$G${last_row},MATCH($A{row_idx},Tabell!$A$2:$A${last_row},0)),\"\")",
        )
        ws_live.cell(
            row=row_idx,
            column=8,
            value=f"=IFERROR(INDEX(Tabell!$H$2:$H${last_row},MATCH($A{row_idx},Tabell!$A$2:$A${last_row},0)),\"\")",
        )
        ws_live.cell(
            row=row_idx,
            column=9,
            value=f"=IFERROR(INDEX(Tabell!$I$2:$I${last_row},MATCH($A{row_idx},Tabell!$A$2:$A${last_row},0)),\"\")",
        )

    for ws in (ws_raw, ws_stats, ws_match, ws_table, ws_live, ws_matches):
        autosize_columns(ws)

    output_path.parent.mkdir(parents=True, exist_ok=True)
    wb.save(output_path)


def main() -> int:
    args = parse_args()
    project_root = Path(__file__).resolve().parents[1]

    dataset_path = Path(args.dataset)
    output_path = Path(args.output)
    matches_path = Path(args.matches_dataset) if args.matches_dataset else None

    if not dataset_path.is_absolute():
        dataset_path = (project_root / dataset_path).resolve()
    if not output_path.is_absolute():
        output_path = (project_root / output_path).resolve()
    if matches_path is not None and not matches_path.is_absolute():
        matches_path = (project_root / matches_path).resolve()

    if not dataset_path.exists():
        print(f"Fant ikke dataset: {dataset_path}")
        return 1

    df = pd.read_csv(dataset_path)

    required = {"Player", "Country", "Goals", "Year"}
    missing = required - set(df.columns)
    if missing:
        print(f"Mangler kolonner i dataset: {sorted(missing)}")
        return 1

    stats_df = build_team_stats(df)
    matches_df = None
    if matches_path is not None:
        if matches_path.exists():
            matches_df = pd.read_csv(matches_path)
        else:
            print(f"Advarsel: Fant ikke kampdataset: {matches_path}")

    create_workbook(df, stats_df, output_path, matches_df)

    print(f"Excel-fil opprettet: {output_path}")
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
