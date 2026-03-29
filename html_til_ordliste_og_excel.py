# Hent ut ord fra HTML-fil (minst 4 bokstaver) og legg til i Excel-kolonne
# Bruk: python html_til_ordliste_og_excel.py --htmlfil sti_til_fil.html --excelfil sti_til_excel.xlsx

import argparse
from bs4 import BeautifulSoup

import argparse
from bs4 import BeautifulSoup
import re
from openpyxl import load_workbook
import requests
import os

parser = argparse.ArgumentParser(description="Trekk ut ord fra HTML og legg til i Excel")
parser.add_argument("--htmlfil", type=str, required=True, help="Sti til HTML-fil eller URL")
parser.add_argument("--excelfil", type=str, required=True, help="Sti til Excel-fil")
parser.add_argument("--ark", type=str, default="Ordliste", help="Ark-navn i Excel")
parser.add_argument("--kolonne", type=str, default="AG", help="Kolonne (f.eks. AG)")
args = parser.parse_args()

# 1. Hent ut ord fra HTML (fra fil eller URL)
if args.htmlfil.startswith("http://") or args.htmlfil.startswith("https://"):
    print(f"Henter HTML fra nettadresse: {args.htmlfil}")
    resp = requests.get(args.htmlfil)
    resp.raise_for_status()
    html_content = resp.text
else:
    print(f"Leser HTML fra fil: {args.htmlfil}")
    with open(args.htmlfil, encoding="utf-8") as f:
        html_content = f.read()

soup = BeautifulSoup(html_content, "html.parser")
tekst = soup.get_text(separator=" ")
# Kun ord med a-z/A-Z, minst 4 bokstaver
ord = re.findall(r"[a-zA-Z]{4,}", tekst)
ordliste = sorted(set(o.lower() for o in ord))
# 2. Åpne Excel og les eksisterende ord
wb = load_workbook(args.excelfil)
if args.ark not in wb.sheetnames:
    print(f"FEIL: Fant ikke ark '{args.ark}' i filen. Tilgjengelige ark: {wb.sheetnames}")
    exit(1)
ws = wb[args.ark]
from openpyxl.utils import column_index_from_string
kol_idx = column_index_from_string(args.kolonne)

eksisterende_ord = set()
for row in ws.iter_rows(min_col=kol_idx, max_col=kol_idx, min_row=1, values_only=True):
    verdi = row[0]
    if verdi:
        eksisterende_ord.add(str(verdi).strip().lower())

# 3. Finn nye ord som ikke finnes fra før
nye_ord = sorted(set(ordliste) - eksisterende_ord)
print(f"Antall nye ord som legges til: {len(nye_ord)}")

# 4. Finn første ledige rad
rad = 1
while ws.cell(row=rad, column=kol_idx).value:
    rad += 1

# 5. Skriv inn nye ord
for o in nye_ord:
    ws.cell(row=rad, column=kol_idx, value=o)
    rad += 1

wb.save(args.excelfil)
print(f"Ferdig! {len(nye_ord)} ord lagt til i {args.excelfil} [{args.ark}!{args.kolonne}]")
