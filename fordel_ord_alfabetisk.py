# Dette skriptet henter alle ord fra kolonne AG i arket "Ordliste" i Ordliste_Norsk_ny.xlsx,
# sorterer dem alfabetisk, og skriver dem ut i kolonnene A:AC (én celle per ord, én rad per bokstav)
# slik at alle ord som starter med samme bokstav havner på samme rad.


print("Starter scriptet...")
from openpyxl import load_workbook
from openpyxl.utils import get_column_letter

EXCEL_FIL = "Ordliste_Norsk_ny.xlsx"
ARK = "Ordliste"
KILDEKOLONNE = "AG"
MALKOLONNER = 29  # A:AC

# 1. Les alle ord fra kolonne AG
def hent_ordliste(fil, ark, kolonne):
    print(f"Åpner fil: {fil}, ark: {ark}, kolonne: {kolonne}")
    wb = load_workbook(fil)
    ws = wb[ark]
    idx = ord(kolonne) - ord('A') + 1 if len(kolonne) == 1 else 33  # AG = 33
    ordliste = []
    for row in ws.iter_rows(min_col=idx, max_col=idx, min_row=1, values_only=True):
        verdi = row[0]
        if verdi and isinstance(verdi, str):
            ordliste.append(verdi.strip())
    print(f"Antall ord hentet: {len(ordliste)}")
    return ordliste, wb, ws

# 2. Sorter ord alfabetisk
ordliste, wb, ws = hent_ordliste(EXCEL_FIL, ARK, KILDEKOLONNE)
print("Sorterer ord...")
ordliste = sorted(set(ordliste), key=lambda x: (x.lower(), x))
print(f"Antall unike ord: {len(ordliste)}")

# 3. Fordel ordene på kolonnene A:AC etter første bokstav
#    (A: ord på A, B: ord på B, ... osv, AA: ord på Æ, AB: ord på Ø, AC: ord på Å)


# Norsk alfabet: A-Z, AA=Æ, AB=Ø, AC=Å
alfabet = [chr(i) for i in range(ord('A'), ord('Z')+1)] + ['Æ', 'Ø', 'Å']
kolonne_map = {bokstav: ("AA" if bokstav=="Æ" else "AB" if bokstav=="Ø" else "AC" if bokstav=="Å" else get_column_letter(ord(bokstav)-ord('A')+1)) for bokstav in alfabet}

# Samle ord per bokstav
ord_per_bokstav = {bokstav: [] for bokstav in alfabet}
for ordet in ordliste:
    if not ordet:
        continue
    for bokstav in alfabet:
        if ordet.upper().startswith(bokstav):
            ord_per_bokstav[bokstav].append(ordet)
            break

# 4. Skriv ut i kolonnene A:Z, AA=Æ, AB=Ø, AC=Å, én rad per ord
max_len = max(len(liste) for liste in ord_per_bokstav.values())

# Skriv ut i kolonnene A:Z, AA=Æ, AB=Ø, AC=Å, én rad per ord

print("Skriver ut ord i kolonner...")
for bokstav in alfabet:
    kol = kolonne_map[bokstav]
    sett = set()
    rad = 1
    for ordet in ord_per_bokstav[bokstav]:
        if ordet not in sett:
            ws[f"{kol}{rad}"] = ordet
            print(f"Legger inn ord '{ordet}' i kolonne {kol}, rad {rad}")
            sett.add(ordet)
            rad += 1
        # Hvis duplikat, hopp over (ikke skriv inn)


# Slett alle verdier i kolonne AG

print("Tømmer kolonne AG...")
from openpyxl.utils import column_index_from_string
ag_idx = column_index_from_string("AG")
for row in ws.iter_rows(min_col=ag_idx, max_col=ag_idx, min_row=1):
    cell = row[0]
    cell.value = None


# Logg dato og antall nye ord i ark 'data'

print("Logger til ark 'data'...")
from datetime import datetime
data_arknavn = "data"
if data_arknavn not in wb.sheetnames:
    wb.create_sheet(data_arknavn)
ws_data = wb[data_arknavn]
# Finn første ledige rad
rad = 1
while ws_data.cell(row=rad, column=1).value:
    rad += 1
ws_data.cell(row=rad, column=1, value=datetime.now().strftime("%Y-%m-%d %H:%M:%S"))
ws_data.cell(row=rad, column=2, value=len(ordliste))

# Tell alle ord i Ordliste!A:AC og logg dette i kolonne 3
print("Teller alle ord i Ordliste!A:AC...")
ord_ark = ws  # ws peker på Ordliste-arket
antall_ord = 0
for col in range(1, 30):  # A:AC = 1 til 29
    # Finn siste rad med innhold i denne kolonnen
    max_row = ord_ark.max_row
    for row in range(1, max_row + 1):
        verdi = ord_ark.cell(row=row, column=col).value
        if verdi is not None and str(verdi).strip() != "":
            antall_ord += 1
ws_data.cell(row=rad, column=3, value=antall_ord)
print(f"Antall ord i Ordliste!A:AC: {antall_ord}")

# Telle tomme og ikke tomme celler i hele arket

## Fjerner treg telling av tomme og fylte celler
# print("Teller tomme og fylte celler...")
# import openpyxl
# empty=0
# filled=0
# path="Ordliste_Norsk_ny.xlsx"
# data_arknavn = "Ordliste"
# if data_arknavn not in wb.sheetnames:
#     wb.create_sheet(data_arknavn)
# ws_data = wb[data_arknavn]
# wb=openpyxl.load_workbook(path)
# sheet=wb.worksheets[0]
# for row in range (sheet.max_row):
#     for column in range (sheet.max_column) :
#         if (sheet.cell(row=row+1, column=column+1).value==""):
#             empty+=1
#         else :
#             filled+=1
# print(f'The number of Empty and Non-Empty cells are {empty} and {filled} respectively. ')


print("Lagrer fil...")
wb.save(EXCEL_FIL)
print(f"Ferdig! Ordene er fordelt alfabetisk i kolonnene A:AC og kolonne AG er tømt i {EXCEL_FIL} (ark: {ARK}). Logg oppdatert i ark 'data'.")
