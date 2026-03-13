import os
import re
import time
import shutil
import sys
from datetime import datetime

import fitz
from openpyxl import Workbook, load_workbook

print("KJØRER HOVEDSCRIPT – 13.03.2026 – 17:00")

# ---------------------------------------------------------
# KONFIGURASJON
# ---------------------------------------------------------
base = r"C:\Strømfakturaer"
innboks = os.path.join(base, "Innboks")
arkiv = os.path.join(base, "Arkiv")
excel_fil = os.path.join(base, "faktura_data.xlsx")

os.makedirs(base, exist_ok=True)
os.makedirs(innboks, exist_ok=True)
os.makedirs(arkiv, exist_ok=True)

# ---------------------------------------------------------
# FUNKSJONER
# ---------------------------------------------------------
def rens_filnavn(filnavn):
    filnavn = filnavn.replace("–", "-").replace("—", "-")
    filnavn = filnavn.replace("’", "'").replace("«", "").replace("»", "")
    filnavn = filnavn.replace("\u00A0", " ")
    filnavn = filnavn.replace("\u00AD", "")
    filnavn = filnavn.encode("ascii", "ignore").decode()
    return filnavn


def vent_til_fil_er_klar(path, timeout=10):
    start = time.time()
    last_size = -1

    while time.time() - start < timeout:
        try:
            size = os.path.getsize(path)
            if size == last_size and size > 0:
                return True
            last_size = size
            time.sleep(0.3)
        except:
            time.sleep(0.3)

    print(f"Advarsel: Filen ble aldri klar: {path}")
    return False


def hent_pdf_tekst(path):
    try:
        doc = fitz.open(path)
    except Exception as e:
        print(f"Kunne ikke åpne PDF: {path} – {e}")
        return ""

    tekst = ""

    for side in doc:
        t1 = side.get_text("text")
        d = side.get_text("dict")
        t2 = ""

        if isinstance(d, dict) and "blocks" in d:
            for b in d["blocks"]:
                if "lines" in b:
                    for line in b["lines"]:
                        for span in line["spans"]:
                            t2 += span.get("text", "") + " "

        blocks = side.get_text("blocks")
        t3 = "\n".join(b[4] for b in blocks if len(b) > 4)

        best = max([t1, t2, t3], key=len)
        tekst += best + "\n"

    tekst = tekst.replace("\u00A0", " ").replace("\u00AD", "")
    return tekst


def try_parse_date(dato_str):
    for fmt in ("%d.%m.%Y", "%d.%m.%y"):
        try:
            return datetime.strptime(dato_str, fmt)
        except:
            pass
    return None


def clean_number(value):
    if not value:
        return value
    return value.replace(" ", "").replace("\xa0", "")


def ensure_sheets(wb):
    if "Fakturaer" not in wb.sheetnames:
        ws_main = wb.create_sheet("Fakturaer")
        ws_main.append(["Filnavn", "Fakturatype", "År-Måned"])
    else:
        ws_main = wb["Fakturaer"]

    if "Wattn" not in wb.sheetnames:
        ws_wattn = wb.create_sheet("Wattn")
        ws_wattn.append([
            "Filnavn", "Fakturanummer", "Fakturadato", "Forfallsdato",
            "Å betale", "KID", "Målepunkt-ID",
            "kWh", "Spotpris", "Påslag",
            "Fastbeløp", "MVA", "Sum inkl mva",
            "Periode", "År-Måned"
        ])
    else:
        ws_wattn = wb["Wattn"]

    if "NVN" not in wb.sheetnames:
        ws_nvn = wb.create_sheet("NVN")
        ws_nvn.append([
            "Filnavn", "Fakturanummer", "Fakturadato", "Forfallsdato",
            "Til gode / Å betale", "Kundenr", "Målepunkt-ID",
            "Energi dag kWh", "Energi dag sum",
            "Energi natt kWh", "Energi natt sum",
            "Kapasitetsledd sum", "Sum nett", "Sum diverse",
            "MVA", "Periode", "År-Måned"
        ])
    else:
        ws_nvn = wb["NVN"]

    return ws_main, ws_wattn, ws_nvn


def sort_sheet_by_date(ws, forfalls_col_index):
    rows = list(ws.iter_rows(min_row=2, values_only=True))

    def sort_key(r):
        dato = r[forfalls_col_index]
        if not dato:
            return datetime.max
        dt = try_parse_date(str(dato))
        return dt if dt else datetime.max

    rows.sort(key=sort_key)

    for i, row in enumerate(rows, start=2):
        for j, value in enumerate(row, start=1):
            ws.cell(row=i, column=j, value=value)


# ---------------------------------------------------------
# TRINN 1: RENS FILNAVN
# ---------------------------------------------------------
nye_filer = []

for fil in os.listdir(innboks):
    full = os.path.join(innboks, fil)
    if not os.path.isfile(full):
        continue

    fil_renset = rens_filnavn(fil)
    ny = os.path.join(innboks, fil_renset)

    if fil_renset != fil:
        try:
            os.rename(full, ny)
            print(f"Renset filnavn: {fil} → {fil_renset}")
        except Exception as e:
            print(f"Kunne ikke rename {fil}: {e}")
            continue

    nye_filer.append(fil_renset)

# ---------------------------------------------------------
# OPPRETT / LAST EXCEL
# ---------------------------------------------------------
if not os.path.exists(excel_fil):
    wb = Workbook()
    std = wb.active
    wb.remove(std)
    ws_main, ws_wattn, ws_nvn = ensure_sheets(wb)
    wb.save(excel_fil)

wb = load_workbook(excel_fil)
ws_main, ws_wattn, ws_nvn = ensure_sheets(wb)

eksisterende = {row[0].value for row in ws_main.iter_rows(min_row=2)}

# ---------------------------------------------------------
# REGEX-PROFILER
# ---------------------------------------------------------
wattn_patterns = {
    "fakturanummer": r"Fakturanr[:\s]*([0-9]+)",
    "fakturadato": r"Fakturadato[:\s]*([0-9.]+)",
    "forfallsdato": r"Forfallsdato[:\s]*([0-9.]+)",
    "a_betale": r"Å betale[:\s]*([0-9.,\s]+)",
    "kid": r"KID-nummer[:\s]*([0-9]+)",
    "malepunkt": r"Målepunkt-ID[:\s]*([0-9]+)",
    "kwh": r"([0-9\s]+,\d+)\s*kWh",
    "spotpris": r"Spot timepris.*?([0-9.,\s]+)\s*øre/kWh",
    "paslag": r"Påslag.*?([0-9.,\s]+)\s*øre/kWh",
    "fastbelop": r"Fastbeløp.*?([0-9.,\s]+)\s*kr/mnd",
    "mva": r"Herav mva\.\s*([0-9.,\s]+)",
    "sum_inkl": r"Sum.*?([0-9.,\s]+)",
    "periode": r"([0-9.]{10}-[0-9.]{10})"
}

nvn_patterns = {
    "fakturanummer": r"Fakturanr\s*([0-9]+)",
    "fakturadato": r"Fakturadato\s*([0-9.]+)",
    "forfallsdato": r"Betalingsfrist\s*([0-9.]+)",
    "a_betale": r"Å betale\s*([0-9.,\s]+)",
    "kundenr": r"Kundenr\s*([0-9]+)",
    "malepunkt": r"Målepunktid\s*([0-9]+)",
    "energi_dag_kwh": r"Energi dag.*?([0-9\s]+,\d+)\s*kWh",
    "energi_dag_sum": r"Energi dag.*?([0-9.,\s]+)\s*$",
    "energi_natt_kwh": r"Energi natt.*?([0-9\s]+,\d+)\s*kWh",
    "energi_natt_sum": r"Energi natt.*?([0-9.,\s]+)\s*$",
    "kapasitetsledd_sum": r"Kapasitetsledd.*?([0-9.,\s]+)\s*$",
    "sum_nett": r"Sum Nett\s*([0-9.,\s]+)",
    "sum_diverse": r"Sum Diverse.*?([\-0-9.,\s]+)",
    "mva": r"Herav mva.*?([0-9.,\s]+)",
    "periode": r"([0-9.]{8}\s*-\s*[0-9.]{8})"
}

tallfelt_som_skal_renses = {
    "a_betale", "kwh", "spotpris", "paslag", "fastbelop",
    "mva", "sum_inkl",
    "energi_dag_kwh", "energi_dag_sum",
    "energi_natt_kwh", "energi_natt_sum",
    "kapasitetsledd_sum", "sum_nett", "sum_diverse"
}

# ---------------------------------------------------------
# TRINN 2: BEHANDLE PDF-FILER
# ---------------------------------------------------------
for fil_renset in nye_filer:

    if fil_renset in eksisterende:
        continue

    pdf_path = os.path.join(innboks, fil_renset)

    if not fil_renset.lower().endswith(".pdf"):
        continue

    if not vent_til_fil_er_klar(pdf_path):
        print(f"Hopper over låst fil: {fil_renset}")
        continue

    tekst = hent_pdf_tekst(pdf_path)

    if not tekst.strip():
        print(f"PDF uten tekst: {fil_renset}")
        continue

    fakturatype = None
    patterns = None

    # Sjekk NVN først
    if any(x in tekst for x in [
        "NORDVEST NETT", "Nordvest Nett", "Nordvestnett", "NVN",
        "Netteigar", "Nettleige", "Energi dag", "Energi natt",
        "Kapasitetsledd", "Sum Nett", "Målarnummer", "Målepunktid",
        "Avrekning"
    ]):
        fakturatype = "NVN"
        patterns = nvn_patterns

    elif "Wattn" in tekst:
        fakturatype = "Wattn"
        patterns = wattn_patterns

    if fakturatype is None:
        print(f"Ukjent fakturatype: {fil_renset}")
        continue

    # Ekstraher data
    data = {}
    for key, pattern in patterns.items():
        m = re.search(pattern, tekst, re.DOTALL)
        verdi = m.group(1).strip() if m else ""
        if key in tallfelt_som_skal_renses and verdi:
            verdi = clean_number(verdi)
        data[key] = verdi

    aar_maned = ""
    if data.get("fakturadato"):
        dt = try_parse_date(data["fakturadato"])
        if dt:
            aar_maned = dt.strftime("%Y-%m")

    # Skriv til riktig ark
    if fakturatype == "Wattn":
        ws_wattn.append([
            fil_renset,
            data.get("fakturanummer", ""), data.get("fakturadato", ""), data.get("forfallsdato", ""),
            data.get("a_betale", ""), data.get("kid", ""), data.get("malepunkt", ""),
            data.get("kwh", ""), data.get("spotpris", ""), data.get("paslag", ""),
            data.get("fastbelop", ""), data.get("mva", ""), data.get("sum_inkl", ""),
            data.get("periode", ""), aar_maned
        ])

    elif fakturatype == "NVN":
        ws_nvn.append([
            fil_renset,
            data.get("fakturanummer", ""), data.get("fakturadato", ""), data.get("forfallsdato", ""),
            data.get("a_betale", ""), data.get("kundenr", ""), data.get("malepunkt", ""),
            data.get("energi_dag_kwh", ""), data.get("energi_dag_sum", ""),
            data.get("energi_natt_kwh", ""), data.get("energi_natt_sum", ""),
            data.get("kapasitetsledd_sum", ""), data.get("sum_nett", ""),
            data.get("sum_diverse", ""), data.get("mva", ""), data.get("periode", ""),
            aar_maned
        ])

    ws_main.append([fil_renset, fakturatype, aar_maned])

    # Flytt PDF til arkiv
    try:
        shutil.move(pdf_path, os.path.join(arkiv, fil_renset))
    except Exception as e:
        print(f"Feil ved flytting av {fil_renset}: {e}")

# ---------------------------------------------------------
# TRINN 3: SORTER FANER
# ---------------------------------------------------------
sort_sheet_by_date(ws_wattn, 3)
sort_sheet_by_date(ws_nvn, 3)

rows_main = list(ws_main.iter_rows(min_row=2, values_only=True))

def sort_key_main(r):
    am = r[2]
    if not am:
        return datetime.max
    try:
        return datetime.strptime(am, "%Y-%m")
    except:
        return datetime.max

rows_main.sort(key=sort_key_main)

for i, row in enumerate(rows_main, start=2):
    for j, value in enumerate(row, start=1):
        ws_main.cell(row=i, column=j, value=value)

# ---------------------------------------------------------
# LAGRE EXCEL
# ---------------------------------------------------------
try:
    wb.save(excel_fil)
except PermissionError:
    print("Kunne ikke lagre Excel-fil. Lukk den og prøv igjen.")
    sys.exit(1)

print("FERDIG! Wattn og NVN er lagt i egne ark, renset, sortert og lagret.")