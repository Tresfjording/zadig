from openpyxl import load_workbook
import string

EXCEL_FIL = "Ordliste_Norsk_ny.xlsx"
ARK = "Ordliste"

wb = load_workbook(EXCEL_FIL, data_only=False)
ws = wb[ARK]

mistenkelige = []
for col in range(1, 30):  # A:AC = 1 til 29
    for row in range(1, ws.max_row + 1):
        cell = ws.cell(row=row, column=col)
        verdi = cell.value
        # Sjekk for mistenkelige celler
        if verdi is not None:
            s = str(verdi)
            # Celler med bare usynlige tegn (men ikke tomme)
            if s.strip() == "" and s != "":
                mistenkelige.append((cell.coordinate, "bare usynlige tegn"))
            # Celler med formler
            if cell.data_type == 'f':
                mistenkelige.append((cell.coordinate, f"formel: {s}"))
            # Celler med spesialtegn
            if any(ord(c) < 32 and c not in '\t\n\r' for c in s):
                mistenkelige.append((cell.coordinate, f"spesialtegn: {repr(s)}"))

print(f"Fant {len(mistenkelige)} mistenkelige celler i Ordliste!A:AC:")
for coord, info in mistenkelige:
    print(f"{coord}: {info}")
if not mistenkelige:
    print("Ingen mistenkelige celler funnet!")
