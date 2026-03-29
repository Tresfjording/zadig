from __future__ import annotations

import argparse
import sys
import tempfile
from pathlib import Path

from openpyxl import load_workbook
from openpyxl.utils import column_index_from_string


SHEET_CANDIDATES = ["Ordliste", "AlleOrd", "Kolonner"]
DEFAULT_EXCEL = "G-Ordliste.xlsm"
DEFAULT_TEXT = "ordliste_unik.txt"
COLUMN_START = "A"
COLUMN_END = "AC"


def normalize_word(value: object) -> str:
    if value is None:
        return ""
    text = str(value).strip().upper()
    return " ".join(text.split())


def resolve_sheet(workbook):
    normalized_map = {name.strip().casefold(): name for name in workbook.sheetnames}
    for candidate in SHEET_CANDIDATES:
        key = candidate.strip().casefold()
        if key in normalized_map:
            actual_name = normalized_map[key]
            return workbook[actual_name], actual_name

    wanted = ", ".join(SHEET_CANDIDATES)
    available = ", ".join(workbook.sheetnames)
    raise KeyError(f"Fant ikke ordliste-ark. Provde: {wanted}. Tilgjengelige ark: {available}")


def iter_text_words(text_path: Path):
    with text_path.open("r", encoding="utf-8") as handle:
        for line in handle:
            word = normalize_word(line)
            if word:
                yield word


def read_grid(sheet, min_col: int, max_col: int) -> list[list[object]]:
    data: list[list[object]] = []
    for row in sheet.iter_rows(min_col=min_col, max_col=max_col, values_only=True):
        data.append(list(row))
    return data


def collect_existing_words(grid: list[list[object]]) -> set[str]:
    words: set[str] = set()
    for row in grid:
        for value in row:
            word = normalize_word(value)
            if word:
                words.add(word)
    return words


def flatten_unique_words(grid: list[list[object]]) -> list[str]:
    seen: set[str] = set()
    flattened: list[str] = []
    for row in grid:
        for value in row:
            word = normalize_word(value)
            if word and word not in seen:
                seen.add(word)
                flattened.append(word)
    return flattened


def reshape_words(words: list[str], width: int) -> list[list[str | None]]:
    rows: list[list[str | None]] = []
    for index in range(0, len(words), width):
        chunk = words[index:index + width]
        row = chunk + [None] * (width - len(chunk))
        rows.append(row)
    if not rows:
        rows.append([None] * width)
    return rows


def rewrite_sheet(sheet, min_col: int, max_col: int, rows: list[list[str | None]]) -> None:
    width = max_col - min_col + 1
    required_rows = len(rows)
    current_rows = sheet.max_row

    for row_index in range(1, max(required_rows, current_rows) + 1):
        row_values = rows[row_index - 1] if row_index <= required_rows else [None] * width
        for offset, value in enumerate(row_values, start=0):
            sheet.cell(row=row_index, column=min_col + offset, value=value)


def save_workbook_atomic(workbook, destination: Path) -> None:
    with tempfile.NamedTemporaryFile(delete=False, suffix=destination.suffix, dir=destination.parent) as handle:
        temp_path = Path(handle.name)

    try:
        workbook.save(temp_path)
        temp_path.replace(destination)
    except PermissionError as exc:
        if temp_path.exists():
            temp_path.unlink()
        raise PermissionError(
            f"Ingen tilgang til å overskrive '{destination}'. Lukk Excel-filen hvis den er åpen, og prov igjen."
        ) from exc
    except Exception:
        if temp_path.exists():
            temp_path.unlink()
        raise


def import_words(text_path: Path, excel_path: Path, output_path: Path | None = None) -> dict[str, int | str]:
    if not text_path.is_file():
        raise FileNotFoundError(f"Fant ikke tekstfilen: {text_path}")
    if not excel_path.is_file():
        raise FileNotFoundError(f"Fant ikke Excel-filen: {excel_path}")

    destination = output_path or excel_path

    workbook = load_workbook(excel_path, keep_vba=True)
    try:
        sheet, actual_sheet_name = resolve_sheet(workbook)
        min_col = column_index_from_string(COLUMN_START)
        max_col = column_index_from_string(COLUMN_END)
        width = max_col - min_col + 1

        grid = read_grid(sheet, min_col=min_col, max_col=max_col)
        existing_words = collect_existing_words(grid)

        input_total = 0
        duplicate_in_input = 0
        new_words: list[str] = []
        staged_new_words: set[str] = set()

        for word in iter_text_words(text_path):
            input_total += 1
            if word in existing_words or word in staged_new_words:
                duplicate_in_input += 1
                continue
            staged_new_words.add(word)
            new_words.append(word)

        all_unique_words = flatten_unique_words(grid)
        before_dedup_count = sum(1 for row in grid for value in row if normalize_word(value))
        all_unique_words.extend(new_words)

        reshaped = reshape_words(all_unique_words, width)
        rewrite_sheet(sheet, min_col=min_col, max_col=max_col, rows=reshaped)
        save_workbook_atomic(workbook, destination)

        after_unique_count = len(all_unique_words)
        removed_duplicates = before_dedup_count + len(new_words) - after_unique_count

        return {
            "sheet": actual_sheet_name,
            "saved_to": str(destination),
            "input_total": input_total,
            "new_words": len(new_words),
            "skipped_existing_or_duplicate": duplicate_in_input,
            "removed_duplicates": removed_duplicates,
            "final_unique_total": after_unique_count,
        }
    finally:
        workbook.close()


def build_parser() -> argparse.ArgumentParser:
    parser = argparse.ArgumentParser(
        description="Hent nye ord fra en tekstfil og legg dem inn i G-Ordliste.xlsm fanen Ordliste."
    )
    parser.add_argument(
        "--txt",
        type=Path,
        default=Path(DEFAULT_TEXT),
        help=f"Tekstfil med ett ord per linje. Standard: {DEFAULT_TEXT}",
    )
    parser.add_argument(
        "--excel",
        type=Path,
        default=Path(DEFAULT_EXCEL),
        help=f"Excel-fil som skal oppdateres. Standard: {DEFAULT_EXCEL}",
    )
    parser.add_argument(
        "--output",
        type=Path,
        default=None,
        help="Valgfri utfil. Bruk denne hvis originalfilen er apen eller du vil lagre til en kopi.",
    )
    return parser


def main() -> int:
    parser = build_parser()
    args = parser.parse_args()

    try:
        result = import_words(text_path=args.txt, excel_path=args.excel, output_path=args.output)
    except Exception as exc:
        print(f"Feil: {exc}", file=sys.stderr)
        return 1

    print(f"Ark: {result['sheet']}")
    print(f"Lagret til: {result['saved_to']}")
    print(f"Linjer lest fra txt: {result['input_total']}")
    print(f"Nye ord lagt til: {result['new_words']}")
    print(f"Hoppet over eksisterende/duplikate ord: {result['skipped_existing_or_duplicate']}")
    print(f"Duplikater fjernet i hele arket: {result['removed_duplicates']}")
    print(f"Totalt antall unike ord i arket: {result['final_unique_total']}")
    return 0


if __name__ == "__main__":
    raise SystemExit(main())